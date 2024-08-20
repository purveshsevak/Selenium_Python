import openpyxl

excelData = {}

book = openpyxl.load_workbook("/Users/purvesh.sevak/Desktop/PythonExcel.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)

sheet.cell(row=4, column=1).value = "TestCase3"

print(sheet.cell(row=4, column=1).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['C2'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase1":
        for j in range(2, sheet.max_column + 1):
            # print(sheet.cell(row=i, column=j).value)
            excelData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print("-------------------------")

print(excelData)
