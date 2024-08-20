import openpyxl


class HomePageData:

    test_userInformation = [{"userName": "Purvesh Sevak", "email": "purveshsevak@gmail.com", "password": "asdf@123",
                             "gender": "Male"}, {"userName": "Test User", "email": "testuser@gmail.com",
                                                 "password": "qwerty@123", "gender": "Female"}]

    @staticmethod
    def getExcelData(test_case_number):
        excelData = {}
        book = openpyxl.load_workbook("/Users/purvesh.sevak/Desktop/PythonExcel.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_number:
                for j in range(2, sheet.max_column + 1):
                    excelData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [excelData]
